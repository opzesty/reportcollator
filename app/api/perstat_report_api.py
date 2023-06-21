from flask import Blueprint, request, jsonify, Response
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter
import xlrd
import sys
import os
import io
from datetime import date

api_blueprint = Blueprint('papi', __name__)

@api_blueprint.route('/generate-perstat', methods=['POST'])
def generate_perstat():
    # Get the request data
    unit = request.form['unit']
    am_pm = request.form['am_pm']
    output_filename = request.form['output_filename']
    perstat_file = request.files['perstat_file']

    # Perform the report generation logic
    time = set_time(am_pm)
    date = set_date()
    zulutime = combine_datetime(time, date)
    location = "Fort McCoy, Wisconsin"
    save_path = get_save_path(output_filename)
    workbook = initialize_workbook()
    workbook = create_headers(workbook)
    perstat_info = pull_perstat_info(perstat_file)
    workbook = populate_report(workbook, perstat_info, unit, zulutime, location)
    workbook = format_report(workbook)

    # Prepare the response
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    # Prepare the response headers
    headers = {
        'Content-Disposition': f'attachment; filename="{output_filename}"',
        'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    }

    return Response(output, headers=headers)

def get_save_path(filename):
    current_directory = os.getcwd()
    save_path = os.path.join(current_directory, filename)
    return save_path

def check_and_add_unit(sheet, search_unit):
    column_a = sheet["A"]

    # Check if the search string is already present in column A
    for row in range(2, len(column_a) + 1):
        if column_a[row - 1].value == search_unit:
            return row

    # If the search string is not found, add it to the end of column A
    new_row = len(column_a) + 1
    sheet.cell(row=new_row, column=1).value = search_unit

    return new_row
    
def set_time(am_pm):
    if am_pm == "am":
        return "0700"
    elif am_pm == "pm":
        return "1300"
    else:
        raise ValueError("Invalid value for am_pm. Expected 'am' or 'pm'.")
    
def set_date():
    current_date = date.today()
    return current_date
    
def combine_datetime(time, current_date):
    formatted_date = current_date.strftime("%d").upper()
    formatted_time = time[:2] + time[2:]
    formatted_month_year = current_date.strftime("%b%y").upper()
    combined_datetime = f"{formatted_date}{formatted_time}Z{formatted_month_year}"
    return combined_datetime

def initialize_workbook():
    # Create a new workbook
    workbook = Workbook()
    
    return workbook
    
def save_workbook(workbook, save_path):
    workbook.save(save_path)
    
    return(save_path)
    
def pull_perstat_info(perstat_file):
    perstat_workbook = xlrd.open_workbook(file_contents=perstat_file.read())
    perstat_sheet = perstat_workbook.sheet_by_index(0)  # Assuming it's the first sheet

    data = {}

    # Iterate over the rows, starting from the second row (index 1)
    for row in range(1, perstat_sheet.nrows):
        rank = perstat_sheet.cell_value(row,1)
        data[rank] = (perstat_sheet.cell_value(row,3), perstat_sheet.cell_value(row,4), perstat_sheet.cell_value(row,5))
    
    return data

def create_headers(workbook):
    # Headers for perstat report
    headers = [
    "Date/Time (DD01300ZJUN23)",
    "LOCATION",
    "UNIT",
    "GRADE",
    "RANK`",
    "BRANCH/MOS",
    "ON HAND",
    "AUTHORIZED",
    "REPLACEMENTS",
    "RETURNED TO DUTY",
    "KILLED",
    "WOUNDED",
    "NON-BATTLE LOSS",
    "MISSING",
    "DESERTERS",
    "AWOL",
    "CAPTURED",
    "AUTHENTICATION"
    ]


    # Remove the default first sheet
    default_sheet = workbook.active
    column= 1
    
    for header in headers:
        unit_cell = default_sheet.cell(row=1, column=column)
        unit_cell.value = header
        unit_cell.font = Font(bold=True, underline="single")
        column+= 1

    return workbook

def populate_report(workbook, data, unit, zulutime, location):
    sheet = workbook.active
    row = 2
    
    rank_dict = {
        "E1": "PVT",
        "E2": "PV2",
        "E3": "PFC",
        "E4": "SPC",
        "E5": "SGT",
        "E6": "SSG",
        "E7": "SFC",
        "E8": "MSG",
        "E9": "SGM",
        "O1": "2LT",
        "O2": "1LT",
        "O3": "CPT",
        "O4": "MAJ",
        "O5": "LTC",
        "O6": "COL",
        "O7": "BG",
        "O8": "MG",
        "O9": "LTG",
        "O10": "GEN"
    }
    
    for rank in data.keys():
        # Set date time group
        cell = sheet.cell(row=row, column = 1)
        cell.value = zulutime
        # Set location
        cell = sheet.cell(row=row, column = 2)
        cell.value = location
        # Set unit
        cell = sheet.cell(row=row, column = 3)
        cell.value = unit
        # Set Rank (number form)
        cell = sheet.cell(row=row, column = 4)
        cell.value = rank
        # Set Rank (common parlance)
        cell = sheet.cell(row=row, column = 5)
        cell.value = rank_dict[rank]
        # Set Branch
        cell = sheet.cell(row=row, column = 6)
        cell.value = data[rank][0]
        # Set On Hand
        cell = sheet.cell(row=row, column = 7)
        cell.value = data[rank][1]
        # Set Authorized
        cell = sheet.cell(row=row, column = 8)
        cell.value = data[rank][2]
        row+=1
        
                
    return workbook
    
def format_report(workbook):
    # Assume active sheet
    sheet = workbook.active
    
    # Create border style
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))
    
    # Iterate over all cells in the sheet
    for row in sheet.iter_rows():
        for cell in row:
            # Apply border style to each cell
            cell.border = border
            
            # Autosize the column to fit the text
            column_letter = get_column_letter(cell.column)
            sheet.column_dimensions[column_letter].auto_size = True
            
    return workbook