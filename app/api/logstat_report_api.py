from flask import Blueprint, request, jsonify, Response
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter
import xlrd
import sys
import os
import io
from datetime import date

api_blueprint = Blueprint('api', __name__)

@api_blueprint.route('/generate-logstat', methods=['POST'])
def generate_logstat():
    # Get the request data
    unit = request.form['unit']
    output_filename = request.form['output_filename']
    am_pm = request.form['am_pm']
    supp_file = request.files['supp_file']
    equip_file = request.files['equip_file']
    print(request.form)

    # Perform the report generation logic
    time = set_time(am_pm)
    date = set_date()
    save_path = get_save_path(output_filename)
    workbook = initialize_workbook()
    workbook = initial_population(workbook)
    supp_information = pull_supp_info(supp_file)
    equip_information = pull_eqp_info(equip_file)
    workbook = populate_report(workbook, supp_information, equip_information, unit, date.strftime('%Y-%m-%d'), time)
    workbook = format_report(workbook)

    # Prepare the response
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    # Prepare the response headers
    headers = {
        'Content-Disposition': f'attachment; filename="{output_filename}"',
        'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    }

    return Response(output, headers=headers)

def get_save_path(filename):
    current_directory = os.getcwd()
    save_path = os.path.join(current_directory, filename)
    return save_path
    
def check_and_add_unit(sheet, search_unit):
    column_a = sheet["A"]

    # Check if the search string is already present in column A
    for row in range(2, len(column_a) + 1):
        if column_a[row - 1].value == search_unit:
            return row

    # If the search string is not found, add it to the end of column A
    new_row = len(column_a) + 1
    sheet.cell(row=new_row, column=1).value = search_unit

    return new_row
    
def set_time(am_pm):
    if am_pm == "am":
        return "0700"
    elif am_pm == "pm":
        return "1300"
    else:
        raise ValueError("Invalid value for am_pm. Expected 'am' or 'pm'.")
    
def set_date():
    current_date = date.today()
    return current_date
    
def combine_datetime(time, current_date):
    formatted_date = current_date.strftime("%d").upper()
    formatted_time = time[:2] + time[2:]
    formatted_month_year = current_date.strftime("%b%y").upper()
    combined_datetime = f"{formatted_date}{formatted_time}Z{formatted_month_year}"
    return combined_datetime

def initialize_workbook():
    # Create a new workbook
    workbook = Workbook()
    
    return workbook
    
def save_workbook(workbook, save_path):
    workbook.save(save_path)
    
    return(save_path)
    
def pull_supp_info(supp_file):
    # Open the Excel spreadsheet
    supp_workbook = xlrd.open_workbook(file_contents=supp_file.read())
    supp_sheet = supp_workbook.sheet_by_index(0)  # Assuming it's the first sheet

    # Initialize the data structure
    data = {}

    # Iterate over the rows, starting from the second row (index 1)
    for row in range(1, supp_sheet.nrows):
        column_h_value = supp_sheet.cell_value(row, 7)  # Column H value
        if "CLASS"  in column_h_value:
            column_c_value = supp_sheet.cell_value(row, 2)  # Column C value
            column_d_value = int(supp_sheet.cell_value(row, 3))  # Column D value
            column_e_value = int(supp_sheet.cell_value(row, 4))  # Column E value

            # Check if the Column H value is already a key in the dictionary
            if column_h_value in data:
                # Append the tuple to the existing list
                data[column_h_value].append((column_c_value, column_d_value, column_e_value))
            else:
                # Create a new list with the tuple
                data[column_h_value] = [(column_c_value, column_d_value, column_e_value)]

    return data
    
def pull_eqp_info(equip_file):
    # Open the second Excel sheet using xlrd
    equip_workbook = xlrd.open_workbook(file_contents=equip_file.stream.read())
    equip_sheet = equip_workbook.sheet_by_index(0)

    # Initialize an empty list to store the matched values
    matched_values = []

    # Iterate over the rows, starting from the second row (index 1)
    for row in range(1, equip_sheet.nrows):
        column_e_value = equip_sheet.cell_value(row, 4)  # Column E value
        column_f_value = int(equip_sheet.cell_value(row, 5))  # Column F value
        column_g_value = int(equip_sheet.cell_value(row, 6))  # Column G value

        # Append the matched values as a tuple to the list
        matched_values.append((column_e_value, column_f_value, column_g_value))

    return matched_values


def initial_population(workbook):
    # Remove the default first sheet
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    
    #  Create 10 sheets
    for sheet_num in range(0, 11):
        # Create a new sheet
        sheet = workbook.create_sheet(title=f"LINE   {sheet_num}")

        # Set "UNIT" column header and formatting
        unit_cell = sheet.cell(row=1, column=1)
        unit_cell.value = "UNIT"
        unit_cell.font = Font(bold=True, underline="single")

    return workbook
    
def update_headers_and_values(sheet, report, this_unit):
    # Iterate over the values in the report
    for value in report:
        header = value[0]
        OH = value[1]
        AUTH = value[2]

        # Check if the header already exists in the first row
        col_num = 1
        header_exists = False
        for cell in sheet[1]:
            if cell.value == header:
                header_exists = True
                break
            col_num += 1

        # If the header doesn't exist, create it in the first row
        if not header_exists:
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, underline="single")

        # Add OH/AUTH values in the second row of the corresponding column
        report_line = "{column_f_value}OH/{column_g_value}AUTH"
        cell = sheet.cell(row=this_unit, column=col_num)
        cell.value = report_line.format(column_f_value=OH, column_g_value=AUTH)

def populate_report(workbook, data, equip, unit, date, time):
    # Iterate over 11 sheets
    for sheet_num in range(0, 11):
        sheet_name = workbook.sheetnames[sheet_num]
        sheet = workbook[sheet_name]
        this_unit = check_and_add_unit(sheet, unit)
        # Only LINE Specific headings
        report = []
        headers = []
        if sheet_num == 0:
            report = [("DATE",""), ("TIME",""), ("LOCATION",""), ("HEADCOUNT","")]
            col_num = 2
            for item in report:
                cell = sheet.cell(row=1,column=col_num)
                cell.value = item[0]
                cell.font = Font(bold=True, underline="single")
                
                cell = sheet.cell(row=this_unit,column=col_num)
                cell.value = item[1]

                date_cell = sheet.cell(row=2, column=2)
                time_cell = sheet.cell(row=2, column=3)
                date_cell.value = date
                time_cell.value = time
                col_num += 1
        elif sheet_num == 1:
            report = data.get("CLASS_I", [])
        elif sheet_num == 2:
            report = data.get("CLASS_II", [])
        elif sheet_num == 3:
            report.extend(data.get("CLASS_III(B)", []))
            report.extend(data.get("CLASS_III(P)", []))
        elif sheet_num == 4:
            report = data.get("CLASS_IV", [])
        elif sheet_num == 5:
            report = data.get("CLASS_V", [])
        elif sheet_num == 7:
            sheet.cell(row=this_unit, column=1).value=unit
            # Iterate over the values and check if a header already exists
            item_number = 1
            for value in equip:
                header_exists = False
                for col_num, cell in enumerate(sheet[1], start=2):
                    if cell.value == value[0]:
                        header_exists = True
                        sheet.cell(row=this_unit, column=col_num).value = "{column_f_value}OH/{column_g_value}AUTH".format(column_f_value=value[1], column_g_value=value[2])
                        item_number += 1
                        break

                if not header_exists:
                    # Find the next empty column and set the value as the header
                    column_index = len(sheet[1]) + 1
                    header_cell = sheet.cell(row=1, column=column_index)
                    header_cell.value = value[0]
                    header_cell.font = Font(bold=True, underline="single")
                    sheet.cell(row=this_unit, column=column_index).value = "{column_f_value}OH/{column_g_value}AUTH".format(column_f_value=value[1], column_g_value=value[2])
                    item_number += 1
        elif sheet_num == 8:
            report = data.get("CLASS_VII", [])
        elif sheet_num == 9:
            report = data.get("CLASS_VIII", [])
        elif sheet_num == 10:
            report = data.get("CLASS_IX", [])
            OH = [value[1] for value in data.get("CLASS_I", [])]
        elif sheet_num == 11:
            report = data.get("CLASS_X", [])
        else:
            report = []

        if sheet_num != 0:
            update_headers_and_values(sheet, report, this_unit)
            
    return workbook
    
        
def format_report(workbook):
    # Iterate over 11 sheets
    for sheet_num in range(0, 11):
        sheet_name = workbook.sheetnames[sheet_num]
        sheet = workbook[sheet_name]
        
        # Create border style
        border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))
        
        # Iterate over all cells in the sheet
        for row in sheet.iter_rows():
            for cell in row:
                # Apply border style to each cell
                cell.border = border
                
                # Autosize the column to fit the text
                column_letter = get_column_letter(cell.column)
                sheet.column_dimensions[column_letter].auto_size = True
                
    return workbook