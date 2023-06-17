from openpyxl import Workbook
from openpyxl.styles import Font
import xlrd
import sys
import os


def get_save_path(filename):
    current_directory = os.getcwd()
    save_path = os.path.join(current_directory, filename)
    return save_path

def identify_unit():
    return sys.argv[1]
    
def check_and_add_unit(sheet, search_unit):
    column_a = sheet["A"]

    # Check if the search string is already present in column A
    for row in range(2, len(column_a) + 1):
        if column_a[row - 1].value == search_unit:
            return row

    # If the search string is not found, add it to the end of column A
    new_row = len(column_a) + 1
    sheet.cell(row=new_row, column=1).value = search_unit

    return new_row
    
def set_time():
    return "0800"
    
def set_date():
    return "06/15/2023"

def initialize_workbook():
    # Create a new workbook
    workbook = Workbook()
    
    return workbook
    
def save_workbook(workbook, save_path):
    workbook.save(save_path)
    
    return(save_path)
    
def pull_supp_info(path):
    # Open the Excel spreadsheet
    SUPP = xlrd.open_workbook(path)
    SUPPsheet = SUPP.sheet_by_index(0)  # Assuming it's the first sheet

    # Initialize the data structure
    data = {}

    # Iterate over the rows, starting from the second row (index 1)
    for row in range(1, SUPPsheet.nrows):
        column_h_value = SUPPsheet.cell_value(row, 7)  # Column H value
        if "CLASS"  in column_h_value:
            column_c_value = SUPPsheet.cell_value(row, 2)  # Column C value
            column_d_value = int(SUPPsheet.cell_value(row, 3))  # Column D value
            column_e_value = int(SUPPsheet.cell_value(row, 4))  # Column E value

            # Check if the Column H value is already a key in the dictionary
            if column_h_value in data:
                # Append the tuple to the existing list
                data[column_h_value].append((column_c_value, column_d_value, column_e_value))
            else:
                # Create a new list with the tuple
                data[column_h_value] = [(column_c_value, column_d_value, column_e_value)]

    return data
    
def pull_eqp_info(path):
    # Open the second Excel sheet using xlrd
    second_workbook = xlrd.open_workbook(filename=path)
    second_sheet = second_workbook.sheet_by_index(0)

    # Initialize an empty list to store the matched values
    matched_values = []

    # Iterate over the rows, starting from the second row (index 1)
    for row in range(1, second_sheet.nrows):
        column_e_value = second_sheet.cell_value(row, 4)  # Column E value
        column_f_value = int(second_sheet.cell_value(row, 5))  # Column F value
        column_g_value = int(second_sheet.cell_value(row, 6))  # Column G value

        # Append the matched values as a tuple to the list
        matched_values.append((column_e_value, column_f_value, column_g_value))
        
    print(matched_values)

    return matched_values


def format_report(workbook):
    # Remove the default first sheet
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    
    #  Create 10 sheets
    for sheet_num in range(0, 11):
        # Create a new sheet
        sheet = workbook.create_sheet(title=f"LINE   {sheet_num}")

        # Set "UNIT" column header and formatting
        unit_cell = sheet.cell(row=1, column=1)
        unit_cell.value = "UNIT"
        unit_cell.font = Font(bold=True, underline="single")

    return workbook
    
def update_headers_and_values(sheet, report, this_unit):
    # Iterate over the values in the report
    for value in report:
        header = value[0]
        OH = value[1]
        AUTH = value[2]

        # Check if the header already exists in the first row
        col_num = 1
        header_exists = False
        for cell in sheet[1]:
            if cell.value == header:
                header_exists = True
                break
            col_num += 1

        # If the header doesn't exist, create it in the first row
        if not header_exists:
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, underline="single")

        # Add OH/AUTH values in the second row of the corresponding column
        report_line = "{column_f_value}OH/{column_g_value}AUTH"
        cell = sheet.cell(row=this_unit, column=col_num)
        cell.value = report_line.format(column_f_value=OH, column_g_value=AUTH)

def populate_report(workbook, data, equip, unit):
    # Iterate over 11 sheets
    for sheet_num in range(0, 11):
        sheet_name = workbook.sheetnames[sheet_num]
        sheet = workbook[sheet_name]
        this_unit = check_and_add_unit(sheet, unit)
        # Only LINE Specific headings
        report = []
        headers = []
        if sheet_num == 0:
            report = [("DATE",""), ("TIME",""), ("LOCATION",""), ("HEADCOUNT","")]
            col_num = 2
            for item in report:
                cell = sheet.cell(row=1,column=col_num)
                cell.value = item[0]
                cell.font = Font(bold=True, underline="single")
                
                cell = sheet.cell(row=this_unit,column=col_num)
                cell.value = item[1]
                col_num += 1
        elif sheet_num == 1:
            report = data.get("CLASS_I", [])
        elif sheet_num == 2:
            report = data.get("CLASS_II", [])
        elif sheet_num == 3:
            report.extend(data.get("CLASS_III(B)", []))
            report.extend(data.get("CLASS_III(P)", []))
        elif sheet_num == 4:
            report = data.get("CLASS_IV", [])
        elif sheet_num == 5:
            report = data.get("CLASS_V", [])
        elif sheet_num == 7:
            sheet.cell(row=this_unit, column=1).value=unit
            # Iterate over the values and check if a header already exists
            item_number = 1
            for value in equip:
                header_exists = False
                for col_num, cell in enumerate(sheet[1], start=2):
                    if cell.value == value[0]:
                        header_exists = True
                        sheet.cell(row=this_unit, column=col_num).value = "{column_f_value}OH/{column_g_value}AUTH".format(column_f_value=value[1], column_g_value=value[2])
                        item_number += 1
                        break

                if not header_exists:
                    # Find the next empty column and set the value as the header
                    column_index = len(sheet[1]) + 1
                    header_cell = sheet.cell(row=1, column=column_index)
                    header_cell.value = value[0]
                    header_cell.font = Font(bold=True, underline="single")
                    sheet.cell(row=this_unit, column=column_index).value = "{column_f_value}OH/{column_g_value}AUTH".format(column_f_value=value[1], column_g_value=value[2])
                    item_number += 1
        elif sheet_num == 8:
            report = data.get("CLASS_VII", [])
        elif sheet_num == 9:
            report = data.get("CLASS_VIII", [])
        elif sheet_num == 10:
            report = data.get("CLASS_IX", [])
            OH = [value[1] for value in data.get("CLASS_I", [])]
        elif sheet_num == 11:
            report = data.get("CLASS_X", [])
        else:
            report = []

        if sheet_num != 0:
            update_headers_and_values(sheet, report, this_unit)
            
    return workbook


if __name__ == "__main__":
    unit = identify_unit()
    time = set_time()
    date = set_date()
    
    
    output_filename = sys.argv[2]

    save_path = get_save_path(output_filename)
    
    workbook = initialize_workbook()
    workbook = format_report(workbook)
    
    supp_sheet = get_save_path(sys.argv[3])
    equip_sheet = get_save_path(sys.argv[4])
    
    supp_information = pull_supp_info(supp_sheet)
    equip_information = pull_eqp_info(equip_sheet)
    
    workbook = populate_report(workbook, supp_information, equip_information, unit)
    
    save_path = save_workbook(workbook, output_filename)